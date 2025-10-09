"""
Data Provider System for SchwabLiveData

This module provides an abstraction layer for market data providers.
- ExcelLiveProvider: Real-time data from Excel via xlwings (RTD)
- SchwabProvider: Placeholder for real Schwab API integration (future)

The provider system allows easy switching between data sources without
changing the rest of the application.
"""

import os
import time
from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional
from decimal import Decimal
from datetime import datetime, timezone
from openpyxl import load_workbook
import threading

# Import Excel Live provider from dedicated module
try:
    from .excel_live import ExcelLiveProvider  # type: ignore
except Exception:
    ExcelLiveProvider = None  # Optional dependency; factory will raise if requested


class BaseProvider(ABC):
    """Abstract base class for all market data providers."""
    
    @abstractmethod
    def get_latest_quotes(self, symbols: List[str]) -> List[Dict[str, Any]]:
        """
        Return latest market data for the given symbols.
        
        Args:
            symbols: List of instrument symbols (e.g., ['/YM', '/ES', '/NQ'])
            
        Returns:
            List of market data dictionaries matching the API schema
        """
        raise NotImplementedError()
    
    @abstractmethod
    def health_check(self) -> Dict[str, Any]:
        """
        Return provider health status and metadata.
        
        Returns:
            Dictionary with provider status, connection info, rate limits, etc.
        """
        raise NotImplementedError()
    
    @abstractmethod
    def get_provider_name(self) -> str:
        """Return the name of this provider."""
        raise NotImplementedError()


class SchwabProvider(BaseProvider):
    """
    Placeholder provider for Schwab API integration.
    
    This will be implemented when Schwab API credentials are available.
    For now, it raises NotImplementedError to clearly indicate it's not ready.
    """
    
    def __init__(self, config: Dict[str, Any]):
        from .schwab_client import SchwabApiClient  # Lazy import to keep dependency surface small
        self.config = config or {}
        self.client = SchwabApiClient(self.config)
        self._connected = False
        # Note: actual OAuth handshake will be performed when permissions are granted
    
    def get_latest_quotes(self, symbols: List[str]) -> List[Dict[str, Any]]:
        """
        Fetch real market data from Schwab API.
        
        TODO: Implement when Schwab API access is available:
        1. Authenticate with OAuth2
        2. Make REST API calls to Schwab
        3. Map Schwab response fields to our standard schema
        4. Handle rate limits and errors
        """
        # The implementation will:
        # - Ensure valid access token (refresh if needed)
        # - Call quotes endpoint and transform response to our schema
        # For now, we raise 501 to signal upstream that the provider exists but is not yet enabled.
        raise NotImplementedError("Schwab API integration pending approval; provider scaffolded but not active.")
    
    def health_check(self) -> Dict[str, Any]:
        """Return health status for Schwab provider."""
        info = self.client.health()
        return {
            "provider": "schwab",
            "connected": info.get("configured", False) and self._connected,
            "status": info.get("status", "not_configured"),
            "auth": info.get("auth", {}),
            "base_url": self.config.get("base_url"),
            "scopes": self.config.get("scopes"),
        }
    
    def get_provider_name(self) -> str:
        return "Schwab API Provider (Not Implemented)"


def create_provider(provider_type: str, **kwargs) -> BaseProvider:
    """
    Factory function to create the appropriate provider.
    
    Args:
        provider_type: 'excel', 'excel_live', or 'schwab'
        **kwargs: Provider-specific configuration
        
    Returns:
        Configured provider instance
        
    Raises:
        ValueError: If provider_type is not supported
    """
    if provider_type.lower() == 'excel':
        excel_file = kwargs.get('excel_file')
        sheet_name = kwargs.get('sheet_name', 'Futures')
        return ExcelProvider(excel_file, sheet_name=sheet_name)
    
    elif provider_type.lower() == 'excel_live':
        if ExcelLiveProvider is None:
            raise RuntimeError("Excel Live provider requested but xlwings is not installed.")
        excel_file = kwargs.get('excel_file')
        sheet_name = kwargs.get('sheet_name', 'Futures')
        live_range = kwargs.get('live_range', 'A1:M16')
        poll_ms = int(kwargs.get('poll_ms', 150))
        require_open = bool(kwargs.get('require_open', False))
        prov = ExcelLiveProvider(
            file_path=excel_file,
            sheet_name=sheet_name,
            range_address=live_range,
            poll_ms=poll_ms,
            require_open=require_open,
        )
        # Start polling thread immediately
        if hasattr(prov, 'start'):
            prov.start()
        return prov
    
    elif provider_type.lower() == 'schwab':
        config = kwargs.get('config', {})
        return SchwabProvider(config)
    
    else:
        raise ValueError(f"Unknown provider type: {provider_type}")


class ExcelProvider(BaseProvider):
    """Provider that reads futures data from an Excel file (xlsx/xlsm)."""

    def __init__(self, excel_file_path: str, sheet_name: str = 'Futures'):
        if not excel_file_path:
            raise ValueError("ExcelProvider requires 'excel_file_path'")
        self.excel_file_path = excel_file_path
        self.sheet_name = sheet_name
        self._last_update = None
        self._iteration_count = 0
        self._file_mtime = os.path.getmtime(self.excel_file_path) if os.path.exists(self.excel_file_path) else None
        self._base_rows = self._load_excel()

    def _load_excel(self) -> List[Dict[str, Any]]:
        try:
            wb = load_workbook(filename=self.excel_file_path, data_only=True, read_only=True)
            if self.sheet_name not in wb.sheetnames:
                # If sheet not found, use the first sheet
                ws = wb[wb.sheetnames[0]]
            else:
                ws = wb[self.sheet_name]

            # Expect a header row
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter)
            headers = [str(h).strip().lower() if h is not None else '' for h in headers]
            # If the first column header is blank, assume it's the symbol column
            if headers and headers[0] == '':
                headers[0] = 'symbol'

            data_rows = []
            for r in rows_iter:
                row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                # Normalize keys used by our schema
                symbol = (row.get('symbol') or row.get('ticker') or row.get('sym'))
                if not symbol:
                    continue
                data_rows.append(row)
            return data_rows
        except FileNotFoundError:
            return []
        except Exception as e:
            raise RuntimeError(f"Error loading Excel file {self.excel_file_path}: {e}")

    def _coalesce(self, row: Dict[str, Any], keys: List[str], default=None):
        for k in keys:
            if k in row and row[k] is not None and row[k] != '':
                return row[k]
        return default

    def get_latest_quotes(self, symbols: List[str]) -> List[Dict[str, Any]]:
        # Helper to parse numbers that may be in fractional-tick format like 116'27 (== 116 + 27/32)
        def parse_numeric_value(v: Any) -> Optional[float]:
            if v is None:
                return None
            if isinstance(v, (int, float, Decimal)):
                try:
                    return float(v)
                except Exception:
                    return None
            s = str(v).strip()
            if s == '':
                return None
            # Remove commas
            s2 = s.replace(',', '')
            # Try simple float
            try:
                return float(s2)
            except Exception:
                pass
            # Try 32nds format: 116'27 or 116'27.5
            import re
            m = re.match(r"^(\d+)'(\d+(?:\.\d+)?)$", s2)
            if m:
                whole = float(m.group(1))
                frac = float(m.group(2)) / 32.0
                return whole + frac
            return None
        # Reload if the file changed to pick up fresh RTD-calculated values
        try:
            mtime = os.path.getmtime(self.excel_file_path)
            if self._file_mtime != mtime:
                self._base_rows = self._load_excel()
                self._file_mtime = mtime
        except Exception:
            pass

        self._iteration_count += 1
        now = datetime.now(timezone.utc).isoformat()
        result: List[Dict[str, Any]] = []

        # Build a map by symbol for quick lookup with aliases
        by_symbol: Dict[str, Dict[str, Any]] = {}
        alias_map = {
            # Common alternate tickers
            'RT': 'RTY',
            'RTY': 'RTY',
            '30YBOND': 'ZB',
            'ZB': 'ZB',
            'YM': '/YM', '/YM': '/YM',
            'ES': '/ES', '/ES': '/ES',
            'NQ': '/NQ', '/NQ': '/NQ',
        }
        for row in self._base_rows:
            sym_raw = self._coalesce(row, ['symbol', 'ticker', 'sym'])
            if sym_raw is None:
                continue
            sym = str(sym_raw).strip()
            if not sym:
                continue
            # Primary mapping as-is
            by_symbol[sym] = row
            # Add with/without leading slash variants for equity index futures
            if sym.startswith('/'):
                by_symbol[sym[1:]] = row
            else:
                by_symbol[f'/{sym}'] = row
            # Add alias if present
            alias = alias_map.get(sym.upper())
            if alias:
                by_symbol[alias] = row

        # Symbol-specific precision defaults (matches futures contract conventions)
        precision_defaults = {
            '/YM': 0, 'YM': 0,  # Dow mini trades in whole points
            '/ES': 2, 'ES': 2,  # Quarter point increments
            '/NQ': 2, 'NQ': 2,  # Quarter point increments
            'RTY': 2,           # Tenth increments
            'CL': 2,            # Penny increments
            'SI': 3,            # Half-cent increments (0.005)
            'HG': 4,            # 0.0005 tick size
            'GC': 1,            # Dime increments
            'VX': 2,            # 0.01
            'DX': 2,            # 0.01
            'ZB': 2,            # 1/32 simplified to 2 decimals
        }

        for symbol in symbols:
            row = by_symbol.get(symbol, {})
            name = self._coalesce(row, ['name', 'description'], symbol)
            exchange = self._coalesce(row, ['exchange', 'exch'], 'CME')
            # Use symbol-specific default or fallback to 2
            default_precision = precision_defaults.get(symbol, 2)
            display_precision = int(self._coalesce(row, ['display_precision', 'precision', 'dp'], default_precision))

            # Pull numeric fields (support various header names)
            def f(keys, default=None):
                v = self._coalesce(row, keys, None)
                val = parse_numeric_value(v)
                return val if val is not None else default

            last = f(['base_price', 'last', 'price'], None)
            bid = f(['bid'], None)
            ask = f(['ask'], None)
            high = f(['high', 'high_price', 'world high', 'whigh'], None)
            low = f(['low', 'low_price', 'world low', 'wlow'], None)
            open_price = f(['open', 'open_price'], None)
            prev_close = f(['previous_close', 'prev_close', 'close_prev', 'close'], None)
            vwap = f(['vwap'], None)
            # Volume can be integer; still try numeric parse
            volume = f(['volume', 'vol'], None)
            volume = int(volume) if volume is not None else None
            signal = self._coalesce(row, ['signal'], 'HOLD')
            # Extended numeric fields should always be numeric to satisfy enrichment step
            stat_value = f(['stat_value', 'stat'], 0.0)
            contract_weight = f(['contract_weight', 'weight'], 1.0)
            change_provided = f(['change', 'num', 'netchange', 'net change'], None)
            change_pct_provided = f(['change_percent', 'perc', 'change%'], None)
            bid_size = f(['bid_size', 'bidsize', 'bid size'], None)
            ask_size = f(['ask_size', 'asksize', 'ask size'], None)
            bid_size = int(bid_size) if bid_size is not None else None
            ask_size = int(ask_size) if ask_size is not None else None

            # Use only provided values from the sheet
            # Calculate vwap if missing but we have bid/ask/last
            if vwap is None and bid is not None and ask is not None and last is not None:
                vwap = (bid + ask + last) / 3

            # Calculate change if not provided
            if change_provided is not None:
                change = change_provided
            else:
                change = (last - prev_close) if (last is not None and prev_close is not None) else None

            # Calculate change percentage if not provided
            if change_pct_provided is not None:
                change_pct = change_pct_provided
            else:
                change_pct = ((change / prev_close) * 100) if (change is not None and prev_close not in (None, 0)) else None

            result.append({
                "instrument": {
                    "id": hash(symbol) % 10000,
                    "symbol": symbol,
                    "name": name,
                    "exchange": exchange,
                    "currency": "USD",
                    "display_precision": display_precision,
                    "is_active": True,
                    "sort_order": symbols.index(symbol),
                },
                "price": str(round(last, display_precision)) if last is not None else None,
                "bid": str(round(bid, display_precision)) if bid is not None else None,
                "ask": str(round(ask, display_precision)) if ask is not None else None,
                "last_size": None,
                "bid_size": bid_size,
                "ask_size": ask_size,
                "open_price": str(open_price) if open_price is not None else None,
                "high_price": str(high) if high is not None else None,
                "low_price": str(low) if low is not None else None,
                "close_price": None,
                "previous_close": str(prev_close) if prev_close is not None else None,
                "change": str(round(change, 4)) if change is not None else None,
                "change_percent": str(round(change_pct, 4)) if change_pct is not None else None,
                "vwap": str(round(vwap, display_precision)) if vwap is not None else None,
                "volume": volume,
                "market_status": "OPEN",
                "data_source": "excel_provider",
                "is_real_time": False,  # No simulation, just real data
                "delay_minutes": 0,
                "extended_data": {
                    "signal": signal,
                    "stat_value": str(stat_value),
                    "contract_weight": str(contract_weight),
                },
                "timestamp": now,
            })

        self._last_update = now
        return result

    def health_check(self) -> Dict[str, Any]:
        exists = os.path.exists(self.excel_file_path)
        return {
            "provider": "excel",
            "connected": exists,
            "excel_file": self.excel_file_path,
            "sheet": self.sheet_name,
            "last_update": self._last_update,
            "iteration_count": self._iteration_count,
        }

    def get_provider_name(self) -> str:
        return "Excel Provider"
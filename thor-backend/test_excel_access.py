#!/usr/bin/env python
"""
Quick test script to check Excel file access
"""
import os
from openpyxl import load_workbook

excel_file = r'A:\Thor\CleanData.xlsm'

print(f"Testing Excel file access: {excel_file}")
print(f"File exists: {os.path.exists(excel_file)}")

if os.path.exists(excel_file):
    try:
        # Test with openpyxl (regular Excel provider)
        wb = load_workbook(excel_file, data_only=True, read_only=True)
        print(f"Available sheets: {wb.sheetnames}")
        
        if 'Futures' in wb.sheetnames:
            ws = wb['Futures']
            print("Sample data from A1:M5:")
            for row in ws['A1:M5']:
                values = [str(cell.value) if cell.value is not None else '' for cell in row]
                print(f"  {values}")
        else:
            print("Sheet 'Futures' not found!")
            
    except Exception as e:
        print(f"Error reading Excel file: {e}")

    # Test xlwings if available
    try:
        import xlwings as xw
        print("\nTesting xlwings (Excel Live):")
        
        # Check if Excel is running
        apps_count = getattr(xw.apps, 'count', 0)
        print(f"Excel applications running: {apps_count}")
        
        if apps_count > 0:
            app = xw.apps.active
            print(f"Open books: {[b.name for b in app.books]}")
            
            # Try to open/connect to the file
            try:
                book = xw.Book(excel_file)
                print(f"Successfully connected to: {book.name}")
                if 'Futures' in [s.name for s in book.sheets]:
                    sheet = book.sheets['Futures']
                    sample_data = sheet.range('A1:M3').value
                    print("Sample data via xlwings:")
                    for row in sample_data[:3]:
                        print(f"  {row}")
                else:
                    print("Sheet 'Futures' not found in xlwings!")
                    
            except Exception as e:
                print(f"xlwings connection error: {e}")
        else:
            print("No Excel applications running - xlwings can't connect")
            
    except ImportError:
        print("xlwings not available")
    except Exception as e:
        print(f"xlwings test error: {e}")

else:
    print("Excel file not found! Please check the path.")
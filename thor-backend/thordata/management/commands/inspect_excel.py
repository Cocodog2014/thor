import argparse
import csv
import json
import os
from pathlib import Path
from typing import Iterable, List, Tuple

from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


def slugify_header(h: str) -> str:
    if h is None:
        return ""
    s = str(h).strip()
    # collapse spaces and slashes to underscores
    for ch in ["/", "\\", "-", " "]:
        s = s.replace(ch, "_")
    # remove duplicate underscores
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def infer_header_row(ws) -> Tuple[int, List[str]]:
    """Find the first non-empty row and use it as headers."""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [str(v).strip() if v is not None else "" for v in row]
        if any(v for v in values):
            headers = [slugify_header(v) for v in values]
            return i, headers
    raise CommandError("No non-empty rows found in sheet")


def sample_rows(ws, start_row: int, max_rows: int = 10, width: int | None = None) -> List[List]:
    rows: List[List] = []
    for i, row in enumerate(ws.iter_rows(min_row=start_row + 1, values_only=True), start=1):
        values = list(row)
        if width is not None:
            values = values[:width]
        rows.append(values)
        if i >= max_rows:
            break
    return rows


def write_csv(ws, headers: List[str], start_row: int, csv_path: Path) -> int:
    """Write rows to CSV, skipping trailing all-empty rows. Returns number of data rows written."""
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    written = 0
    empty_streak = 0
    # Threshold to decide we've reached the true end of data
    EMPTY_BREAK_THRESHOLD = 100
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in ws.iter_rows(min_row=start_row + 1, values_only=True):
            values = list(row[: len(headers)])
            is_empty = True
            for v in values:
                if v is None:
                    continue
                if isinstance(v, str) and v.strip() == "":
                    continue
                # Any non-empty value marks the row as non-empty
                is_empty = False
                break
            if is_empty:
                empty_streak += 1
                if empty_streak >= EMPTY_BREAK_THRESHOLD:
                    break
                continue
            # reset streak and write the row
            empty_streak = 0
            writer.writerow(["" if v is None else v for v in values])
            written += 1
    return written


class Command(BaseCommand):
    help = "Inspect an Excel/CSV file: report sheets, header row, sample data; optionally export to CSV."

    def add_arguments(self, parser: argparse.ArgumentParser) -> None:
        parser.add_argument("path", type=str, help="Path to .xlsx/.xlsm or .csv file")
        parser.add_argument("--sheet", type=str, help="Excel sheet name (for .xlsx)")
        parser.add_argument("--max-rows", type=int, default=10, help="Sample rows to print")
        parser.add_argument("--export-csv", action="store_true", help="Export normalized CSV (defaults next to source unless --out-dir provided)")
        parser.add_argument("--out-dir", type=str, help="Directory to write exported CSV into (used with --export-csv)")
        parser.add_argument("--header-row", type=int, help="1-based header row index override (for .xlsx)")
        parser.add_argument("--no-trim-trailing-cols", action="store_true", help="Do not trim trailing empty columns (default trims)")
        parser.add_argument("--trim-lookahead", type=int, default=200, help="Rows to scan when trimming trailing empty columns")

    @staticmethod
    def _is_empty_value(v) -> bool:
        if v is None:
            return True
        if isinstance(v, str) and v.strip() == "":
            return True
        return False

    def _determine_effective_width(self, headers: List[str], ws, start_row: int, lookahead: int = 200) -> int:
        """Compute effective width by trimming trailing columns with empty headers and empty data in lookahead rows."""
        raw_width = len(headers)
        # last non-empty header position
        last_header_idx = -1
        for idx in range(raw_width - 1, -1, -1):
            if not self._is_empty_value(headers[idx]):
                last_header_idx = idx
                break
        width = last_header_idx + 1 if last_header_idx >= 0 else 0
        if width < raw_width and lookahead > 0:
            max_row = start_row + lookahead
            for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, values_only=True):
                limit = min(raw_width, len(row))
                for idx in range(limit - 1, -1, -1):
                    if not self._is_empty_value(row[idx]):
                        if idx + 1 > width:
                            width = idx + 1
                        break
        return max(width, 0)

    def handle(self, *args, **options):
        path = Path(options["path"]).expanduser()
        sheet = options.get("sheet")
        max_rows = options["max_rows"]
        export_csv = options["export_csv"]
        out_dir = options.get("out_dir")
        header_row_override = options.get("header_row")
        no_trim_cols = options.get("no_trim_trailing_cols", False)
        trim_lookahead = int(options.get("trim_lookahead", 200))

        if not path.exists():
            raise CommandError(f"File not found: {path}")
        ext = path.suffix.lower()
        if ext == ".csv":
            # Simple CSV profile
            with path.open("r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                headers = [slugify_header(h) for h in headers]
                rows = []
                for i, row in enumerate(reader, start=1):
                    rows.append(row)
                    if i >= max_rows:
                        break
            self.stdout.write(json.dumps({
                "type": "csv",
                "path": str(path),
                "columns": len(headers),
                "headers": headers,
                "sample_rows": rows,
            }, indent=2, default=str))
            return

        if ext not in (".xlsx", ".xlsm"):
            raise CommandError("Unsupported file type. Provide .xlsx, .xlsm, or .csv")

        if openpyxl is None:
            raise CommandError("openpyxl not installed. Run pip install openpyxl")

        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        if sheet and sheet not in sheet_names:
            raise CommandError(f"Sheet '{sheet}' not found. Available: {sheet_names}")

        ws = wb[sheet] if sheet else wb.active
        if header_row_override and header_row_override > 0:
            cells = next(ws.iter_rows(min_row=header_row_override, max_row=header_row_override, values_only=True))
            headers = [slugify_header(c) for c in cells]
            start_row = header_row_override
        else:
            start_row, headers = infer_header_row(ws)
        raw_columns = len(headers)
        if not no_trim_cols:
            eff_width = self._determine_effective_width(headers, ws, start_row=start_row, lookahead=trim_lookahead)
            if eff_width < raw_columns:
                headers = headers[:eff_width]
        else:
            eff_width = raw_columns

        rows = sample_rows(ws, start_row=start_row, max_rows=max_rows, width=eff_width)

        info = {
            "type": "xlsx",
            "path": str(path),
            "sheet": ws.title,
            "raw_columns": raw_columns,
            "columns": len(headers),
            "headers": headers,
            "header_row_index": start_row,
            "sample_rows": rows,
        }

        # Optional CSV export
        if export_csv:
            if out_dir:
                target_dir = Path(out_dir).expanduser()
                target_dir.mkdir(parents=True, exist_ok=True)
                csv_path = target_dir / f"{path.stem}-{ws.title}.csv"
            else:
                csv_path = path.with_suffix("")
                csv_path = csv_path.parent / f"{path.stem}-{ws.title}.csv"
            written_rows = write_csv(ws, headers, start_row=start_row, csv_path=csv_path)
            info["export_csv_path"] = str(csv_path)
            info["exported_data_rows"] = written_rows

        self.stdout.write(json.dumps(info, indent=2, default=str))
